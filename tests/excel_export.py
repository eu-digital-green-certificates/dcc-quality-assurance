import os 
import json
import logging
import openpyxl
from glob import glob
from pathlib import Path
from argparse import ArgumentParser

config = {
    # Can be overridden by placing a config.json file in the directory
    "base_url" : "https://github.com/eu-digital-green-certificates/dcc-quality-assurance/blob/main/",
    "column_titles" : ["Issuing Country", "Schema Version", "Certificate Type", "Validation Status", "Code URL", "Filename"],
    "column_value_ids" : ["country", "version", "type", None, "url", "file" ],
    "sheet" : "Codes",

    "__countryfile-doc__" : "Following section can be omitted when not using country files feature",
    "countryfile-sheet" : "Validation Results", "countryfile-startrow" : 4,
    "countryfile-ccc" : "G2",
    "countryfile-constants" : {
        "H2" : "Validation Cycle #", 
        "J2" : "Validation Period Text"
    }
}

def main(args):
    workbook = _get_or_create_xlsx(args.filename, config['sheet'])
    workbook[config['sheet']].delete_rows(2, amount=1000)

    file_entry_handlers = [] # List of objects that handle matching files 
    file_entry_handlers.append(lambda entry : _append_row( workbook[config['sheet']], entry ) )

    if args.country_template is not None:
        try: 
            countryFileGenerator = CountryFileGenerator(args.country_template)
            file_entry_handlers.append( lambda entry: countryFileGenerator.addEntry(entry) )
        except: 
            logging.error('Country file template was given but could not be loaded')
            raise
            return -1

    # Main loop: Find all matches and pass them to all handlers
    for directory in _get_country_directories():
        for match in _matching_files(directory):
            for handle in file_entry_handlers:
                handle(match)



    workbook.save(args.filename)
    if args.country_template is not None:
        countryFileGenerator.finalize()

def _append_row(sheet, value_dict):
    values = [ value_dict.get(value_id) for value_id in config['column_value_ids']]
    values = [ value if value is not None else '' for value in values ]
    sheet.append(values)


def _get_or_create_xlsx(filename, sheet_to_use='Codes'):
    try: 
        wb = openpyxl.load_workbook(filename)
    except: 
        wb = openpyxl.Workbook()
        wb.active.title = sheet_to_use
        wb.active.append(config['column_titles'])

    if not sheet_to_use in wb.sheetnames:
        wb.create_sheet(sheet_to_use)
        wb[sheet_to_use].append(config['column_titles'])

    return wb

def _matching_files(directory):
    certificate_types = ['TEST','VAC','REC']
    
    for ctype in certificate_types:
        for match in glob(str(Path(directory,'*' , f'{ctype}*.png'))):
            version = match.split(os.sep)[-2]
            yield { 'type':ctype, 
                    'country':directory, 
                    'version':version, 
                    'url' : config['base_url']+match.replace(os.sep,'/'),
                    'file' : Path(match).name }

    for ctype in certificate_types:
        for match in glob(str(Path(directory, '*' ,'specialcases' , f'{ctype}*.png'))):
            version = match.split(os.sep)[-3]
            yield { 'type':f'{ctype} SpecialCase', 
                    'country':directory, 
                    'version':version, 
                    'url' : config['base_url']+match.replace(os.sep,'/'),
                    'file' : Path(match).name }


def _get_country_directories():
    # A country directory is any directory that has a name of exactly 2 letters
    return [ dirname for dirname in glob('??') if dirname.isalpha() ]

class CountryFileGenerator: 
    '''Generates country files from a template. In order to do so, must first collect 
       reference data from source'''

    def __init__(self, template_file_name):        
        self.countries = set()
        self.template_file_name = template_file_name
        self.wb = openpyxl.load_workbook(template_file_name)
        self.current_row = config["countryfile-startrow"]
        #self.wb[config['countryfile-sheet']].delete_rows(config['countryfile-startrow'], amount=1000)

    def addEntry(self, entry):
        self.countries |= set([entry['country']])
        self.wb = 
        pass

    def finalize(self):
        self.wb.save(self.template_file_name)

if __name__ == '__main__':
    try: 
        import coloredlogs
        coloredlogs.install()
    except:
        pass # If we don't have colored logs, it's not important

    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s')

    try: 
        config = json.load(open('config.json'))
        logging.info('Loaded config.json')
    except:
        logging.info('Using default configuration. Create a config.json if you want to override.')

    parser = ArgumentParser(description='Excel export ')
    parser.add_argument('filename', default='report.xlsx', help='Output file')
    parser.add_argument('--country-template', default=None, help='Generate country files from template')
    args = parser.parse_args()
    main(args)